import base64
import logging
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import io
import zipfile
import openpyxl
import xlrd
_logger = logging.getLogger(__name__)

class ExcelConverterWizard(models.TransientModel):
    _name = 'excel.converter.wizard'
    _description = 'Excel to PDF Converter Wizard'

    attachment_ids = fields.Many2many(
        'ir.attachment', 
        string='Excel Files', 
        required=True,
        help="Upload one or more Excel files to convert to PDF."
    )
    
    search_word = fields.Char(
        string='Search Word',
        help="Enter a word to search within Excel files. Only files containing this word will be included in the filtered ZIP download.",
        default=''
    )

    def action_convert_files(self):
        """Convert all attached Excel files to PDF."""
        self.ensure_one()
        excel_file_obj = self.env['excel.file']
        converted_files_ids = []
        failed_files = []

        if not self.attachment_ids:
            raise UserError(_("Please upload at least one Excel file."))

        for attachment in self.attachment_ids:
            excel_record = None
            try:
                if '.xls' not in attachment.name.lower():
                    _logger.warning(f"Skipping non-Excel file: {attachment.name}")
                    failed_files.append({
                        'name': attachment.name, 
                        'error': 'Not a valid Excel file format (.xls, .xlsx)'
                    })
                    continue

                file_data = attachment.datas
                file_name = attachment.name

                excel_record = excel_file_obj.create({
                    'name': file_name,
                    'excel_file': file_data,
                    'state': 'processing',
                })

                success, pdf_data, pdf_filename, error_msg = excel_record.convert_excel_to_pdf(
                    file_data, file_name
                )
                if success:
                    excel_record.write({
                        'pdf_file': pdf_data,
                        'pdf_filename': pdf_filename,
                        'state': 'converted',
                        'error_message': False,
                    })
                    converted_files_ids.append(excel_record.id)
                else:
                    excel_record.write({
                        'state': 'error',
                        'error_message': error_msg,
                    })
                    failed_files.append({'name': file_name, 'error': error_msg})

            except Exception as e:
                _logger.error(f"Error processing file {attachment.name}: {e}", exc_info=True)
                error_str = str(e)
                failed_files.append({'name': attachment.name, 'error': error_str})
                if excel_record:
                    excel_record.write({'state': 'error', 'error_message': error_str})
                else:
                     excel_file_obj.create({
                        'name': attachment.name,
                        'state': 'error',
                        'error_message': error_str
                    })

        return {
            'type': 'ir.actions.act_window',
            'name': _('Converted Files'),
            'res_model': 'excel.file',
            'view_mode': 'tree,form',
            'domain': [('id', 'in', converted_files_ids)],
            'context': {'search_default_converted': 1},
            'target': 'current',
        }

    def _search_word_in_excel(self, file_data, search_word):
        """Search for a word in Excel file content."""
        if not search_word:
            return True
            
        search_word_lower = search_word.lower().strip()
        _logger.info(f"Searching for word: '{search_word_lower}'")
            
        try:
            # Decode base64 data
            excel_bytes = base64.b64decode(file_data)
            
            # Try to read as xlsx first
            try:
                _logger.info("Trying to read as XLSX format")
                workbook = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    _logger.info(f"Checking sheet: {sheet_name}")
                    sheet = workbook[sheet_name]
                    
                    # Método más eficiente para XLSX
                    for row in sheet.iter_rows(values_only=True):
                        for cell_value in row:
                            if cell_value is not None:
                                # Convertir a string y manejar diferentes tipos de datos
                                try:
                                    cell_str = str(cell_value).lower().strip()
                                    if search_word_lower in cell_str:
                                        _logger.info(f"Found '{search_word}' in XLSX: '{cell_value}'")
                                        workbook.close()
                                        return True
                                except (UnicodeDecodeError, UnicodeEncodeError) as unicode_error:
                                    _logger.warning(f"Unicode error processing cell: {unicode_error}")
                                    continue
                                    
                workbook.close()
                _logger.info("Word not found in XLSX file")
                return False
                
            except Exception as xlsx_error:
                _logger.info(f"XLSX read failed: {xlsx_error}, trying XLS format")
                
                # If xlsx fails, try xls format
                try:
                    workbook = xlrd.open_workbook(file_contents=excel_bytes)
                    
                    for sheet_idx in range(workbook.nsheets):
                        sheet = workbook.sheet_by_index(sheet_idx)
                        sheet_name = workbook.sheet_names()[sheet_idx]
                        _logger.info(f"Checking XLS sheet: {sheet_name}")
                        
                        for row_idx in range(sheet.nrows):
                            for col_idx in range(sheet.ncols):
                                try:
                                    cell_value = sheet.cell_value(row_idx, col_idx)
                                    if cell_value:
                                        cell_str = str(cell_value).lower().strip()
                                        if search_word_lower in cell_str:
                                            _logger.info(f"Found '{search_word}' in XLS cell ({row_idx},{col_idx}): '{cell_value}'")
                                            return True
                                except (UnicodeDecodeError, UnicodeEncodeError) as unicode_error:
                                    _logger.warning(f"Unicode error processing XLS cell: {unicode_error}")
                                    continue
                                    
                    _logger.info("Word not found in XLS file")
                    return False
                    
                except Exception as xls_error:
                    _logger.error(f"XLS read also failed: {xls_error}")
                    return False
                    
        except Exception as e:
            _logger.error(f"Error searching in Excel file: {e}", exc_info=True)
            return False
            
        return False

    def action_download_converted_pdfs_zip(self):
        """Download ZIP with all PDFs or filtered PDFs based on search word."""
        self.ensure_one()

        if not self.attachment_ids:
            raise UserError(_("Please upload at least one Excel file."))

        # Validar y limpiar el término de búsqueda - manejar el caso donde llega como False
        if self.search_word is False or self.search_word is None:
            search_term = ''
        else:
            search_term = str(self.search_word).strip()
        
        use_filter = bool(search_term)
        
        _logger.info(f"DEBUG - Raw search_word field: '{self.search_word}'")
        _logger.info(f"DEBUG - search_word type: {type(self.search_word)}")
        _logger.info(f"DEBUG - Processed search_term: '{search_term}'")
        _logger.info(f"Starting download with filter: {use_filter}, search word: '{search_term}'")
        
        excel_file_obj = self.env['excel.file']
        buffer = io.BytesIO()
        zip_file = zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED)

        errors = []
        files_with_word = []
        files_without_word = []
        processed_count = 0

        for attachment in self.attachment_ids:
            try:
                # Validación más estricta del formato de archivo
                file_name_lower = attachment.name.lower()
                if not (file_name_lower.endswith('.xls') or file_name_lower.endswith('.xlsx')):
                    _logger.warning(f"Skipping non-Excel file: {attachment.name}")
                    errors.append(f"{attachment.name}: Not a valid Excel file (.xls or .xlsx)")
                    continue

                file_data = attachment.datas
                file_name = attachment.name
                
                _logger.info(f"Processing file: {file_name}")

                # Aplicar filtro si existe
                if use_filter:
                    _logger.info(f"Checking if '{file_name}' contains word '{search_term}'")
                    contains_word = self._search_word_in_excel(file_data, search_term)
                    _logger.info(f"File '{file_name}' contains word '{search_term}': {contains_word}")
                    
                    if contains_word:
                        files_with_word.append(file_name)
                        _logger.info(f"Including {file_name} (contains search word)")
                    else:
                        files_without_word.append(file_name)
                        _logger.info(f"Excluding {file_name} (doesn't contain search word)")
                        continue  # Skip this file

                # Convertir archivo a PDF
                excel_record = excel_file_obj.create({
                    'name': file_name,
                    'excel_file': file_data,
                    'state': 'processing',
                })

                success, pdf_data, pdf_filename, error_msg = excel_record.convert_excel_to_pdf(
                    file_data, file_name
                )

                if success and pdf_data:
                    try:
                        # Verificar que pdf_data sea válido antes de decodificar
                        pdf_bytes = base64.b64decode(pdf_data)
                        zip_file.writestr(pdf_filename, pdf_bytes)
                        processed_count += 1
                        
                        excel_record.write({
                            'pdf_file': pdf_data,
                            'pdf_filename': pdf_filename,
                            'state': 'converted',
                        })
                        _logger.info(f"Successfully converted and added {file_name} to ZIP")
                    except Exception as decode_error:
                        _logger.error(f"Error decoding PDF data for {file_name}: {decode_error}")
                        errors.append(f"{file_name}: Error processing PDF data")
                        excel_record.write({
                            'state': 'error',
                            'error_message': f"Error processing PDF data: {decode_error}",
                        })
                else:
                    excel_record.write({
                        'state': 'error',
                        'error_message': error_msg or "Unknown conversion error",
                    })
                    errors.append(f"{file_name}: {error_msg or 'Unknown conversion error'}")
                    _logger.error(f"Failed to convert {file_name}: {error_msg}")

            except Exception as e:
                _logger.error(f"Error converting {attachment.name}: {e}", exc_info=True)
                errors.append(f"{attachment.name}: {str(e)}")

        zip_file.close()

        # Verificar si se procesó algún archivo
        if processed_count == 0:
            if use_filter:
                if not files_with_word and files_without_word:
                    message = _("No files contain the search word '%s'.") % search_term
                    message += _("\n\nFiles checked:\n• %s") % '\n• '.join(files_without_word)
                else:
                    message = _("No PDFs could be generated for files containing the word '%s'.") % search_term
                    if files_without_word:
                        message += _("\n\nFiles that don't contain the word '%s':\n• %s") % (
                            search_term, '\n• '.join(files_without_word)
                        )
                if errors:
                    message += _("\n\nErrors:\n• %s") % '\n• '.join(errors)
            else:
                message = _("No PDFs could be generated.")
                if errors:
                    message += _("\n\nErrors:\n• %s") % '\n• '.join(errors)
            raise UserError(message)

        # Crear el archivo ZIP para descarga
        zip_filename = f"filtered_pdfs_{search_term}.zip" if use_filter else "converted_pdfs.zip"
        _logger.info(f"Created ZIP: {zip_filename} with {processed_count} files")
            
        zip_attachment = self.env['ir.attachment'].create({
            'name': zip_filename,
            'type': 'binary',
            'datas': base64.b64encode(buffer.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/zip',
        })

        download_url = f'/web/content/{zip_attachment.id}?download=true'
        return {
            "type": "ir.actions.act_url",
            "url": download_url,
            "target": "self",
        }